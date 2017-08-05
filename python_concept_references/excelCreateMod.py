import openpyxl, os

os.chdir('c:\\Users\\wffpi\\Documents')

wb = openpyxl.Workbook() # creates new workbook object
print(wb.get_sheet_names()) # new workbook has single sheet names 'Sheet'
sheet = wb.get_sheet_by_name('Sheet')
# to add data
sheet['A1'] = 42
sheet['A2'] = 'Hello'
sheet2 = wb.create_sheet()
sheet2.title = 'New Sheet'
wb.create_sheet(index=0, title='Other Sheet') # creating new sheet at index

wb.save('example.xlsx')

