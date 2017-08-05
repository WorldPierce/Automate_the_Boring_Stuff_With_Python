# install openpyxl to open spreadsheets
import openpyxl, os

os.chdir('c:\\wffpi\\documents')

workbook = openpyxl.load_workbook('example.xlsx')
type(workbook)
sheet = workbook.get_sheet_by_name('Sheet1')
type(sheet)

workbook.get_sheet_names()
cell = sheet['A1'] # gets cell object for cell A 1
cell.value # gives you cell value
str(sheet[A1].value) # prints string value of any cell

sheet.cell(row=1, column=2) # returns cell object same as sheet['B1']
# useful for loops so you don't convert to A, B etc..
for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)
